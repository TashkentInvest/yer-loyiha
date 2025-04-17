import pandas as pd
import re
import requests
from urllib.parse import urlparse, parse_qs
from openpyxl import load_workbook

def extract_coordinates_from_gmaps_url(url):
    """Extract coordinates from a Google Maps URL."""
    if not url or not isinstance(url, str) or "maps.app.goo.gl" not in url:
        return None, None

    try:
        # First, follow the shortened URL to get the actual URL
        response = requests.get(url, allow_redirects=True)
        final_url = response.url

        # Try to extract coordinates from the URL
        # Pattern 1: /@latitude,longitude,
        match = re.search(r'/@(-?\d+\.\d+),(-?\d+\.\d+)', final_url)
        if match:
            return float(match.group(1)), float(match.group(2))

        # Pattern 2: /ll=latitude,longitude
        parsed_url = urlparse(final_url)
        params = parse_qs(parsed_url.query)

        # Check various parameters that might contain coordinates
        for param in ['ll', 'q', 'center']:
            if param in params and ',' in params[param][0]:
                coords = params[param][0].split(',')
                if len(coords) >= 2:
                    try:
                        return float(coords[0]), float(coords[1])
                    except ValueError:
                        pass

        # If no coordinates found in URL params
        return None, None

    except Exception as e:
        print(f"Error processing URL {url}: {e}")
        return None, None

def main():
    # File path
    file_path = r"2024_data_complete_20250416_112034_complete.xlsx"

    # Load the workbook and get the active sheet
    print(f"Loading Excel file: {file_path}")
    wb = load_workbook(file_path)
    sheet = wb.active

    # Find column indices
    header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    geo_col_idx = None
    lat_col_idx = None
    lng_col_idx = None

    for idx, col_name in enumerate(header_row, 1):
        if col_name == "geolocation":
            geo_col_idx = idx
        elif col_name == "lat":
            lat_col_idx = idx
        elif col_name == "lng":
            lng_col_idx = idx

    if not all([geo_col_idx, lat_col_idx, lng_col_idx]):
        print("Could not find all required columns (geolocation, lat, lng)")
        return

    # Process each row
    row_count = sheet.max_row
    processed_count = 0

    print(f"Processing {row_count-1} rows...")
    for row_idx in range(2, row_count + 1):  # Start from row 2 (after header)
        geo_cell = sheet.cell(row=row_idx, column=geo_col_idx)

        if geo_cell.value:
            lat, lng = extract_coordinates_from_gmaps_url(geo_cell.value)

            if lat is not None and lng is not None:
                sheet.cell(row=row_idx, column=lat_col_idx, value=lat)
                sheet.cell(row=row_idx, column=lng_col_idx, value=lng)
                processed_count += 1

    # Save the workbook
    output_file = file_path.replace(".xlsx", "_with_coordinates.xlsx")
    wb.save(output_file)
    print(f"Processing completed. Updated {processed_count} rows with coordinates.")
    print(f"Saved output to: {output_file}")

if __name__ == "__main__":
    main()
